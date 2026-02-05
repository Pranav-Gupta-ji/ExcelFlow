import streamlit as st
import pandas as pd
import os
import logging
import traceback
from datetime import datetime
from openpyxl import load_workbook
import re

# ==================================
# Logging Configuration
# ==================================
LOG_FILE = "app.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)
logger.info("Application started")

# ==================================
# Utility: Check if Excel file is locked (Windows-safe)
# ==================================
def is_file_locked(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "a"):
            pass
        return False
    except PermissionError:
        return True

# ==================================
# Utility: Resolve Safe Sheet Name
# ==================================
def resolve_sheet_name(workbook_path, requested_name):
    if not requested_name or requested_name.strip() == "":
        base_name = "Output"
    else:
        base_name = re.sub(r"[\\/*?:\[\]]", "_", requested_name.strip())

    if not os.path.exists(workbook_path):
        return f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M')}"

    wb = load_workbook(workbook_path)
    existing_sheets = wb.sheetnames

    if base_name not in existing_sheets:
        return base_name

    counter = 1
    while f"{base_name}_{counter}" in existing_sheets:
        counter += 1

    return f"{base_name}_{counter}"

# ==================================
# Streamlit UI
# ==================================
st.set_page_config(page_title="ExcelFlow", layout="wide")
st.title("📊 ExcelFlow")

# ==================================
# Upload Input File
# ==================================
uploaded_file = st.file_uploader(
    "Upload data file (CSV / JSON / Excel)",
    type=["csv", "json", "xlsx"]
)

if uploaded_file:
    try:
        logger.info(f"File uploaded: {uploaded_file.name}")

        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith(".json"):
            df = pd.read_json(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)

        logger.info("Input file loaded successfully")

        st.subheader("📄 Preview of Input Data")
        st.dataframe(df.head())

    except Exception:
        logger.error("Failed to read input file")
        logger.error(traceback.format_exc())
        st.error("❌ Failed to read uploaded file")
        st.stop()

    # ==================================
    # Column Selection
    # ==================================
    st.subheader("🧱 Select Required Columns")
    selected_columns = st.multiselect(
        "Choose columns for output",
        df.columns.tolist()
    )

    if selected_columns:
        logger.info(f"Selected columns: {selected_columns}")

        # ==================================
        # Column Ordering
        # ==================================
        st.subheader("🔀 Arrange Column Order")
        col_order = {}

        for col in selected_columns:
            col_order[col] = st.number_input(
                f"Position for '{col}'",
                min_value=1,
                max_value=len(selected_columns),
                value=selected_columns.index(col) + 1
            )

        ordered_columns = [
            col for col, _ in sorted(col_order.items(), key=lambda x: x[1])
        ]

        logger.info(f"Ordered columns: {ordered_columns}")

        final_df = df[ordered_columns]

        st.subheader("✅ Final Output Preview")
        st.dataframe(final_df.head())

        # ==================================
        # Output Options
        # ==================================
        st.subheader("📁 Output Options")

        output_mode = st.radio(
            "Where do you want to store output?",
            ["Create New Excel Workbook", "Append to Existing Workbook"]
        )

        logger.info(f"Output mode: {output_mode}")

        workbook_path = None

        if output_mode == "Create New Excel Workbook":
            workbook_path = st.text_input(
                "Enter new workbook name",
                value="output.xlsx"
            )
        else:
            existing_file = st.file_uploader(
                "Browse existing Excel workbook",
                type=["xlsx"]
            )
            if existing_file:
                workbook_path = existing_file.name
                logger.info(f"Existing workbook selected: {workbook_path}")

        sheet_name_input = st.text_input(
            "Sheet name (leave blank for auto-generated)"
        )

        # ==================================
        # Save Output (PRODUCTION SAFE)
        # ==================================
        if st.button("💾 Save Output"):
            try:
                if not workbook_path:
                    st.warning("⚠ Please select or enter a workbook")
                    st.stop()

                # ---- FILE LOCK CHECK ----
                if is_file_locked(workbook_path):
                    logger.error(f"Excel file is locked: {workbook_path}")
                    st.error(
                        "❌ The Excel file is currently open or locked.\n\n"
                        "Please close it and try again."
                    )
                    st.stop()

                sheet_name = resolve_sheet_name(
                    workbook_path,
                    sheet_name_input
                )

                logger.info(
                    f"Saving output | File: {workbook_path} | Sheet: {sheet_name}"
                )

                file_exists = os.path.exists(workbook_path)

                if output_mode == "Append to Existing Workbook" and file_exists:
                    with pd.ExcelWriter(
                        workbook_path,
                        engine="openpyxl",
                        mode="a",
                        if_sheet_exists="new"
                    ) as writer:
                        final_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False
                        )
                else:
                    with pd.ExcelWriter(
                        workbook_path,
                        engine="openpyxl",
                        mode="w"
                    ) as writer:
                        final_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False
                        )

                logger.info("Data saved successfully")
                st.success(f"✅ Data saved to sheet '{sheet_name}'")

            except PermissionError:
                logger.error("Permission error while saving Excel")
                logger.error(traceback.format_exc())
                st.error(
                    "❌ Permission denied.\n\n"
                    "Close the Excel file and try again."
                )

            except Exception:
                logger.error("Unexpected error during save")
                logger.error(traceback.format_exc())
                st.error("❌ Failed to save data. Check app.log for details.")
